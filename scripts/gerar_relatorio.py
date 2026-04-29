"""
gerar_relatorio.py
Motor de automação do Pacote de Fechamento Mensal
Rede Máxima Supermercados Ltda.

Uso:
    python scripts/gerar_relatorio.py --mes 2024-06
    python scripts/gerar_relatorio.py --mes 2024-06 --loja L01-Centro
    python scripts/gerar_relatorio.py --mes 2024-06 --formato excel
    python scripts/gerar_relatorio.py --mes 2024-06 --formato pdf
    python scripts/gerar_relatorio.py --mes 2024-06 --formato excel,pdf  (padrão)
"""

from __future__ import annotations

import argparse
import io
import sys
from datetime import datetime
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import pandas as pd

# ── Paths ────────────────────────────────────────────────────────────────────
BASE = Path(__file__).parent.parent
RAW  = BASE / "data" / "raw"
OUT  = BASE / "output"
OUT.mkdir(exist_ok=True)

# ── Paleta de cores (consistente com P1 e P2) ────────────────────────────────
C = {
    "bg":      "#0F1117",
    "surface": "#1A1D27",
    "border":  "#2A2D3E",
    "text":    "#E8EAF0",
    "muted":   "#6B7490",
    "blue":    "#4F8EF7",
    "teal":    "#34D1BF",
    "green":   "#2DD4A0",
    "amber":   "#F5A623",
    "coral":   "#F26B6B",
    "purple":  "#A78BFA",
}
LOJAS_ORDEM = [
    "L01-Centro", "L02-Norte", "L03-Sul", "L04-Leste", "L05-Aeroporto"
]

# ═══════════════════════════════════════════════════════════════════════════════
# ETAPA 2 — ETL e cálculo de KPIs
# ═══════════════════════════════════════════════════════════════════════════════

def carregar_dados() -> dict[str, pd.DataFrame]:
    return {
        "dre":    pd.read_csv(RAW / "dre.csv"),
        "fc":     pd.read_csv(RAW / "fluxo_caixa.csv"),
        "cr":     pd.read_csv(RAW / "contas_receber.csv"),
        "cc":     pd.read_csv(RAW / "centro_custos.csv"),
        "orc":    pd.read_csv(RAW / "orcamento.csv"),
    }


def calcular_kpis(dfs: dict, mes: str, loja: str = "all") -> dict:
    dre = dfs["dre"]
    fc  = dfs["fc"]
    cr  = dfs["cr"]
    cc  = dfs["cc"]

    # Filtro de mês atual e mês anterior
    meses_disp = sorted(dre["mes"].unique())
    idx = meses_disp.index(mes)
    mes_ant = meses_disp[idx - 1] if idx > 0 else None

    def filtrar(df, m, col_loja="loja"):
        d = df[df["mes"] == m]
        if loja != "all" and col_loja in df.columns:
            d = d[d[col_loja] == loja]
        return d

    cur  = filtrar(dre, mes)
    prev = filtrar(dre, mes_ant) if mes_ant else None

    def soma(df, col):
        return df[col].sum() if df is not None and len(df) > 0 else 0

    rl_cur  = soma(cur, "receita_liquida")
    rl_prev = soma(prev, "receita_liquida")
    rb_cur  = soma(cur, "receita_bruta")
    lb_cur  = soma(cur, "lucro_bruto")
    ebitda  = soma(cur, "ebitda")
    ll_cur  = soma(cur, "lucro_liquido")

    mg_bruta   = lb_cur  / rl_cur  if rl_cur  else 0
    mg_ebitda  = ebitda  / rl_cur  if rl_cur  else 0
    mg_liquida = ll_cur  / rl_cur  if rl_cur  else 0

    # Inadimplência
    cr_mes = cr[cr["mes"] == mes]
    cart_total  = cr_mes["total_carteira"].sum()
    cart_vencido = cr_mes["total_vencido"].sum()
    inadim = cart_vencido / cart_total if cart_total else 0

    # Fluxo de caixa
    fc_mes = fc[fc["mes"] == mes]
    fco    = fc_mes["fco"].sum() if len(fc_mes) else 0
    saldo_final = fc_mes["saldo_final"].sum() if len(fc_mes) else 0
    conv_caixa = fc_mes["recebimentos"].sum() / rl_cur if rl_cur else 0

    # Variação receita MoM
    var_receita = (rl_cur - rl_prev) / rl_prev if rl_prev else 0

    # Real vs orçado (DRE)
    rb_orc = soma(cur, "receita_orcada")
    ebitda_orc = soma(cur, "ebitda_orcado")
    var_receita_orc  = (rb_cur - rb_orc) / rb_orc if rb_orc else 0
    var_ebitda_orc   = (ebitda - ebitda_orc) / abs(ebitda_orc) if ebitda_orc else 0

    # Score de saúde financeira (mesmo critério do P1)
    s_mg = min(mg_ebitda / 0.15, 1) * 40
    s_fc = min(max(fco, 0) / (rl_cur * 0.10), 1) * 30 if rl_cur else 0
    s_inadim = max(0, (0.18 - inadim) / 0.18) * 30
    score = round(s_mg + s_fc + s_inadim, 1)

    return {
        "mes": mes,
        "loja": loja,
        "receita_bruta": rb_cur,
        "receita_liquida": rl_cur,
        "lucro_bruto": lb_cur,
        "ebitda": ebitda,
        "lucro_liquido": ll_cur,
        "margem_bruta": mg_bruta,
        "margem_ebitda": mg_ebitda,
        "margem_liquida": mg_liquida,
        "inadimplencia": inadim,
        "fco": fco,
        "saldo_final": saldo_final,
        "conversao_caixa": conv_caixa,
        "var_receita_mom": var_receita,
        "var_receita_orc": var_receita_orc,
        "var_ebitda_orc": var_ebitda_orc,
        "score_saude": score,
        "receita_orcada": rb_orc,
        "ebitda_orcado": ebitda_orc,
    }


def semaforo_ebitda(v: float) -> str:
    return "VERDE" if v >= 0.20 else ("AMARELO" if v >= 0.15 else "VERMELHO")

def semaforo_inadim(v: float) -> str:
    return "VERDE" if v <= 0.10 else ("AMARELO" if v <= 0.18 else "VERMELHO")

def semaforo_score(v: float) -> str:
    return "VERDE" if v >= 70 else ("AMARELO" if v >= 50 else "VERMELHO")


def gerar_comentario(kpis: dict) -> list[str]:
    """Gera comentário executivo automático baseado nos KPIs."""
    linhas = []
    mes_fmt = datetime.strptime(kpis["mes"], "%Y-%m").strftime("%B/%Y").capitalize()
    loja_txt = f"da {kpis['loja']}" if kpis["loja"] != "all" else "consolidado"

    linhas.append(
        f"O desempenho financeiro {loja_txt} em {mes_fmt} apresentou receita bruta de "
        f"R$ {kpis['receita_bruta']:,.0f}, com variação de "
        f"{kpis['var_receita_mom']:+.1%} em relação ao mês anterior."
    )

    sinal_orc = "acima" if kpis["var_receita_orc"] >= 0 else "abaixo"
    linhas.append(
        f"A receita ficou {abs(kpis['var_receita_orc']):.1%} {sinal_orc} do orçado, "
        f"enquanto o EBITDA apresentou desvio de {kpis['var_ebitda_orc']:+.1%} frente à meta."
    )

    sem = semaforo_ebitda(kpis["margem_ebitda"])
    linhas.append(
        f"A margem EBITDA atingiu {kpis['margem_ebitda']:.1%} "
        f"(meta: ≥ 20% | status: {sem}), e a margem bruta ficou em "
        f"{kpis['margem_bruta']:.1%}."
    )

    sem_i = semaforo_inadim(kpis["inadimplencia"])
    linhas.append(
        f"A taxa de inadimplência foi de {kpis['inadimplencia']:.1%} sobre a carteira total "
        f"(meta: ≤ 10% | status: {sem_i}). "
        f"O saldo de caixa encerrou o mês em R$ {kpis['saldo_final']:,.0f}."
    )

    sem_s = semaforo_score(kpis["score_saude"])
    linhas.append(
        f"O Score de Saúde Financeira atingiu {kpis['score_saude']:.0f}/100 "
        f"(status: {sem_s}), composto por margens ({kpis['margem_ebitda']:.1%}), "
        f"geração de caixa (FCO: R$ {kpis['fco']:,.0f}) e nível de inadimplência."
    )

    return linhas


# ═══════════════════════════════════════════════════════════════════════════════
# Helpers de gráfico
# ═══════════════════════════════════════════════════════════════════════════════

def _estilo_dark(fig, ax):
    fig.patch.set_facecolor(C["surface"])
    ax.set_facecolor(C["bg"])
    ax.tick_params(colors=C["muted"], labelsize=8)
    ax.xaxis.label.set_color(C["muted"])
    ax.yaxis.label.set_color(C["muted"])
    for spine in ax.spines.values():
        spine.set_edgecolor(C["border"])
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(
        lambda x, _: f"R$ {x/1e6:.1f}M" if abs(x) >= 1e6 else f"R$ {x/1e3:.0f}K"
    ))


def grafico_receita_ebitda(dfs: dict, mes: str) -> bytes:
    dre = dfs["dre"].copy()
    meses_disp = sorted(dre["mes"].unique())
    idx = meses_disp.index(mes)
    periodo = meses_disp[max(0, idx - 11): idx + 1]
    cons = dre[dre["mes"].isin(periodo)].groupby("mes").agg(
        receita_liquida=("receita_liquida", "sum"),
        ebitda=("ebitda", "sum"),
    ).reset_index()
    labels = [m[5:] for m in cons["mes"]]
    x = np.arange(len(labels))

    fig, ax = plt.subplots(figsize=(7, 2.8))
    _estilo_dark(fig, ax)
    ax.bar(x - 0.18, cons["receita_liquida"], 0.35, color=C["blue"],  alpha=0.85, label="Receita Líq.")
    ax.bar(x + 0.18, cons["ebitda"],           0.35, color=C["teal"],  alpha=0.85, label="EBITDA")
    ax.set_xticks(x); ax.set_xticklabels(labels, rotation=45, ha="right", fontsize=7)
    ax.legend(fontsize=7, labelcolor=C["muted"], framealpha=0)
    ax.set_title("Receita Líquida vs EBITDA — últimos 12 meses", color=C["text"], fontsize=9, pad=6)
    plt.tight_layout()
    buf = io.BytesIO(); fig.savefig(buf, format="png", dpi=130, bbox_inches="tight"); plt.close(fig)
    return buf.getvalue()


def grafico_ranking_lojas(dfs: dict, mes: str) -> bytes:
    sub = dfs["dre"][dfs["dre"]["mes"] == mes].copy()
    sub["margem_ebitda"] = sub["ebitda"] / sub["receita_liquida"]
    sub = sub.sort_values("margem_ebitda", ascending=True)
    cores = [C["green"] if v >= 0.20 else (C["amber"] if v >= 0.15 else C["coral"])
             for v in sub["margem_ebitda"]]

    fig, ax = plt.subplots(figsize=(7, 2.6))
    _estilo_dark(fig, ax)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{x:.0%}"))
    nomes = [l.split("-")[1] for l in sub["loja"]]
    bars = ax.barh(nomes, sub["margem_ebitda"], color=cores, alpha=0.9)
    ax.axvline(0.20, color=C["green"], linewidth=0.8, linestyle="--", alpha=0.6)
    ax.axvline(0.15, color=C["amber"], linewidth=0.8, linestyle="--", alpha=0.6)
    for bar, val in zip(bars, sub["margem_ebitda"]):
        ax.text(val + 0.002, bar.get_y() + bar.get_height() / 2,
                f"{val:.1%}", va="center", color=C["text"], fontsize=8)
    ax.set_title("Ranking de lojas — Margem EBITDA", color=C["text"], fontsize=9, pad=6)
    plt.tight_layout()
    buf = io.BytesIO(); fig.savefig(buf, format="png", dpi=130, bbox_inches="tight"); plt.close(fig)
    return buf.getvalue()


def grafico_real_vs_orcado(dfs: dict, mes: str) -> bytes:
    cur = dfs["dre"][dfs["dre"]["mes"] == mes].groupby("loja").agg(
        receita_bruta=("receita_bruta", "sum"),
        receita_orcada=("receita_orcada", "sum"),
    ).reindex(LOJAS_ORDEM)
    nomes = [l.split("-")[1] for l in cur.index]
    x = np.arange(len(nomes))
    fig, ax = plt.subplots(figsize=(7, 2.8))
    _estilo_dark(fig, ax)
    ax.bar(x - 0.18, cur["receita_bruta"],   0.35, color=C["blue"],   alpha=0.85, label="Real")
    ax.bar(x + 0.18, cur["receita_orcada"],  0.35, color=C["purple"], alpha=0.60, label="Orçado")
    ax.set_xticks(x); ax.set_xticklabels(nomes, fontsize=8)
    ax.legend(fontsize=7, labelcolor=C["muted"], framealpha=0)
    ax.set_title("Real vs Orçado por loja — Receita Bruta", color=C["text"], fontsize=9, pad=6)
    plt.tight_layout()
    buf = io.BytesIO(); fig.savefig(buf, format="png", dpi=130, bbox_inches="tight"); plt.close(fig)
    return buf.getvalue()


def grafico_inadimplencia(dfs: dict, mes: str) -> bytes:
    meses_disp = sorted(dfs["dre"]["mes"].unique())
    idx = meses_disp.index(mes)
    periodo = meses_disp[max(0, idx - 11): idx + 1]
    cr = dfs["cr"][dfs["cr"]["mes"].isin(periodo)].groupby("mes").agg(
        total_vencido=("total_vencido", "sum"),
        total_carteira=("total_carteira", "sum"),
    ).reset_index()
    cr["inadim"] = cr["total_vencido"] / cr["total_carteira"]
    labels = [m[5:] for m in cr["mes"]]
    cores_pts = [C["coral"] if v > 0.18 else (C["amber"] if v > 0.10 else C["green"])
                 for v in cr["inadim"]]

    fig, ax = plt.subplots(figsize=(7, 2.6))
    _estilo_dark(fig, ax)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{x:.0%}"))
    ax.plot(labels, cr["inadim"], color=C["amber"], linewidth=1.5, marker="o",
            markersize=5, markerfacecolor=C["surface"])
    for i, (v, c) in enumerate(zip(cr["inadim"], cores_pts)):
        ax.scatter(i, v, color=c, s=40, zorder=5)
    ax.axhline(0.10, color=C["green"], linewidth=0.8, linestyle="--", alpha=0.6)
    ax.axhline(0.18, color=C["coral"], linewidth=0.8, linestyle="--", alpha=0.6)
    ax.set_xticks(range(len(labels))); ax.set_xticklabels(labels, rotation=45, ha="right", fontsize=7)
    ax.set_title("Taxa de inadimplência — últimos 12 meses", color=C["text"], fontsize=9, pad=6)
    plt.tight_layout()
    buf = io.BytesIO(); fig.savefig(buf, format="png", dpi=130, bbox_inches="tight"); plt.close(fig)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# ETAPA 3 — Gerador Excel (.xlsx)
# ═══════════════════════════════════════════════════════════════════════════════

def gerar_excel(dfs: dict, kpis: dict, mes: str, loja: str, path_out: Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import (Alignment, Border, Font, GradientFill,
                                  PatternFill, Side)
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.formatting.rule import ColorScaleRule, DataBar, DataBarRule

    wb = Workbook()
    wb.remove(wb.active)

    # ── Estilos ──────────────────────────────────────────────────────────────
    def fnt(bold=False, size=10, color="E8EAF0", name="Arial"):
        return Font(bold=bold, size=size, color=color, name=name)
    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)
    def aln(h="center", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    def borda_fina():
        s = Side(style="thin", color="2A2D3E")
        return Border(left=s, right=s, top=s, bottom=s)

    BG_HEADER  = "1A1D27"
    BG_SURFACE = "0F1117"
    BG_ROW1    = "14172A"
    BG_ROW2    = "1A1D27"
    TEXT_MAIN  = "E8EAF0"
    TEXT_MUTED = "6B7490"
    BLUE       = "4F8EF7"
    TEAL       = "34D1BF"
    GREEN      = "2DD4A0"
    AMBER      = "F5A623"
    CORAL      = "F26B6B"
    PURPLE     = "A78BFA"

    def _col(hex6): return hex6.replace("#", "")

    def aplicar_header(ws, linha, colunas: list, bg=BG_HEADER):
        for col, (titulo, w) in enumerate(colunas, 1):
            c = ws.cell(linha, col, titulo)
            c.font = fnt(bold=True, size=9, color=TEXT_MAIN)
            c.fill = fill(bg)
            c.alignment = aln()
            c.border = borda_fina()
            ws.column_dimensions[get_column_letter(col)].width = w

    def formatar_valor(c, fmt="moeda"):
        if fmt == "moeda":
            c.number_format = 'R$ #,##0;(R$ #,##0);"-"'
        elif fmt == "pct":
            c.number_format = '0.0%;(0.0%);"-"'
        elif fmt == "pct2":
            c.number_format = '0.00%'
        elif fmt == "num":
            c.number_format = '#,##0'
        c.alignment = aln("right")

    # ─── ABA 1: Resumo Executivo ─────────────────────────────────────────────
    ws1 = wb.create_sheet("Resumo Executivo")
    ws1.sheet_view.showGridLines = False
    ws1.sheet_properties.tabColor = BLUE

    # Cabeçalho
    ws1.merge_cells("A1:H1")
    t = ws1["A1"]
    t.value = "REDE MÁXIMA SUPERMERCADOS LTDA. — PACOTE DE FECHAMENTO MENSAL"
    t.font = fnt(bold=True, size=13, color=TEXT_MAIN)
    t.fill = fill(BG_HEADER); t.alignment = aln()
    ws1.row_dimensions[1].height = 30

    ws1.merge_cells("A2:H2")
    mes_fmt = datetime.strptime(mes, "%Y-%m").strftime("%B de %Y").upper()
    t2 = ws1["A2"]
    t2.value = f"Competência: {mes_fmt}  |  Gerado em: {datetime.now():%d/%m/%Y %H:%M}"
    t2.font = fnt(size=9, color=TEXT_MUTED); t2.fill = fill(BG_HEADER)
    t2.alignment = aln(); ws1.row_dimensions[2].height = 16

    # KPI cards (linha 4-5)
    kpi_items = [
        ("Receita Bruta", kpis["receita_bruta"], "moeda", None),
        ("Receita Líquida", kpis["receita_liquida"], "moeda", None),
        ("EBITDA", kpis["ebitda"], "moeda", None),
        ("Margem EBITDA", kpis["margem_ebitda"], "pct", semaforo_ebitda(kpis["margem_ebitda"])),
        ("Margem Bruta", kpis["margem_bruta"], "pct", None),
        ("Inadimplência", kpis["inadimplencia"], "pct", semaforo_inadim(kpis["inadimplencia"])),
        ("FCO", kpis["fco"], "moeda", None),
        ("Score Saúde", kpis["score_saude"], "num", semaforo_score(kpis["score_saude"])),
    ]
    for col, (titulo, valor, fmt, sem) in enumerate(kpi_items, 1):
        ws1.column_dimensions[get_column_letter(col)].width = 16
        c_tit = ws1.cell(4, col, titulo)
        c_tit.font = fnt(size=8, color=TEXT_MUTED); c_tit.fill = fill(BG_SURFACE)
        c_tit.alignment = aln(); c_tit.border = borda_fina()
        ws1.row_dimensions[4].height = 18

        c_val = ws1.cell(5, col, valor)
        formatar_valor(c_val, fmt)
        bg_kpi = BG_SURFACE
        if sem == "VERDE":   bg_kpi = "0D2E1E"
        elif sem == "AMARELO": bg_kpi = "2E250D"
        elif sem == "VERMELHO": bg_kpi = "2E0D0D"
        c_val.fill = fill(bg_kpi)
        c_val.font = fnt(bold=True, size=11, color=TEXT_MAIN)
        c_val.alignment = aln(); c_val.border = borda_fina()
        ws1.row_dimensions[5].height = 28

    # Variações (linha 7)
    ws1.merge_cells("A7:H7")
    hd = ws1["A7"]
    hd.value = "Variações e desvios"
    hd.font = fnt(bold=True, size=9, color=TEXT_MAIN); hd.fill = fill(BG_HEADER)
    hd.alignment = aln("left"); ws1.row_dimensions[7].height = 18

    var_items = [
        ("Receita MoM", kpis["var_receita_mom"], "pct"),
        ("Receita vs Orç.", kpis["var_receita_orc"], "pct"),
        ("EBITDA vs Orç.", kpis["var_ebitda_orc"], "pct"),
        ("Conversão Caixa", kpis["conversao_caixa"], "pct"),
        ("Saldo Final Caixa", kpis["saldo_final"], "moeda"),
        ("Lucro Líquido", kpis["lucro_liquido"], "moeda"),
        ("Margem Líquida", kpis["margem_liquida"], "pct"),
        ("Score /100", kpis["score_saude"], "num"),
    ]
    for col, (titulo, valor, fmt) in enumerate(var_items, 1):
        ct = ws1.cell(8, col, titulo)
        ct.font = fnt(size=8, color=TEXT_MUTED); ct.fill = fill(BG_SURFACE)
        ct.alignment = aln(); ct.border = borda_fina(); ws1.row_dimensions[8].height = 18
        cv = ws1.cell(9, col, valor)
        formatar_valor(cv, fmt)
        if fmt == "pct" and titulo not in ("Conversão Caixa",):
            cv.font = fnt(bold=True, size=10,
                          color=GREEN if valor >= 0 else CORAL)
        else:
            cv.font = fnt(bold=True, size=10, color=TEXT_MAIN)
        cv.fill = fill(BG_SURFACE); cv.alignment = aln(); cv.border = borda_fina()
        ws1.row_dimensions[9].height = 24

    # Comentário executivo (linha 11+)
    ws1.merge_cells("A11:H11")
    hc = ws1["A11"]
    hc.value = "Comentário Executivo (gerado automaticamente)"
    hc.font = fnt(bold=True, size=9, color=TEXT_MAIN); hc.fill = fill(BG_HEADER)
    hc.alignment = aln("left"); ws1.row_dimensions[11].height = 18
    for i, linha_txt in enumerate(gerar_comentario(kpis), 12):
        ws1.merge_cells(f"A{i}:H{i}")
        c = ws1.cell(i, 1, linha_txt)
        c.font = fnt(size=9, color=TEXT_MAIN)
        c.fill = fill(BG_ROW1 if i % 2 == 0 else BG_ROW2)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = borda_fina(); ws1.row_dimensions[i].height = 30

    # ─── ABA 2: DRE por Loja ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("DRE por Loja")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = TEAL

    ws2.merge_cells("A1:K1")
    h = ws2["A1"]
    h.value = f"DRE POR LOJA — {mes_fmt}"
    h.font = fnt(bold=True, size=12, color=TEXT_MAIN); h.fill = fill(BG_HEADER)
    h.alignment = aln(); ws2.row_dimensions[1].height = 26

    cols_dre = [
        ("Loja", 18), ("Rec. Bruta", 16), ("Deduções", 14),
        ("Rec. Líquida", 16), ("CMV", 14), ("Lucro Bruto", 16),
        ("Total OPEX", 14), ("EBITDA", 14), ("Mg. Bruta", 11),
        ("Mg. EBITDA", 11), ("Var. vs Orç.", 12),
    ]
    aplicar_header(ws2, 2, cols_dre)

    sub = dfs["dre"][dfs["dre"]["mes"] == mes].copy()
    if loja != "all":
        sub = sub[sub["loja"] == loja]
    sub["mg_bruta"]  = sub["lucro_bruto"] / sub["receita_liquida"]
    sub["mg_ebitda"] = sub["ebitda"] / sub["receita_liquida"]
    sub["var_orc"]   = (sub["receita_bruta"] - sub["receita_orcada"]) / sub["receita_orcada"]
    sub = sub.reindex(sub.index[sub["loja"].isin(LOJAS_ORDEM)])

    for i, (_, row) in enumerate(sub.iterrows(), 3):
        bg = BG_ROW1 if i % 2 == 0 else BG_ROW2
        vals = [
            row["loja"], row["receita_bruta"], row["deducoes"],
            row["receita_liquida"], row["cmv"], row["lucro_bruto"],
            row["total_opex"], row["ebitda"], row["mg_bruta"],
            row["mg_ebitda"], row["var_orc"],
        ]
        fmts = ["txt","moeda","moeda","moeda","moeda","moeda","moeda","moeda","pct","pct","pct"]
        for col, (val, fmt) in enumerate(zip(vals, fmts), 1):
            c = ws2.cell(i, col, val)
            c.fill = fill(bg); c.border = borda_fina()
            if fmt == "moeda":
                formatar_valor(c, "moeda")
                c.font = fnt(size=9, color=TEXT_MAIN)
            elif fmt == "pct":
                c.number_format = "0.0%"
                c.alignment = aln("right")
                if col in (9, 10):
                    c.font = fnt(size=9, color=GREEN if val >= 0.18 else (AMBER if val >= 0.13 else CORAL))
                elif col == 11:
                    c.font = fnt(size=9, color=GREEN if val >= 0 else CORAL)
                else:
                    c.font = fnt(size=9, color=TEXT_MAIN)
            else:
                c.font = fnt(size=9, color=TEXT_MAIN)
                c.alignment = aln("left")
        ws2.row_dimensions[i].height = 18

    # Linha de totais
    ultima = 3 + len(sub)
    ws2.merge_cells(f"A{ultima}:A{ultima}")
    ws2.cell(ultima, 1, "TOTAL").font = fnt(bold=True, size=9, color=TEXT_MAIN)
    ws2.cell(ultima, 1).fill = fill(BG_HEADER); ws2.cell(ultima, 1).border = borda_fina()
    for col_idx, col_let in zip(range(2, 9), ["B","C","D","E","F","G","H"]):
        c = ws2.cell(ultima, col_idx, f"=SUM({col_let}3:{col_let}{ultima-1})")
        formatar_valor(c, "moeda")
        c.font = fnt(bold=True, size=9, color=TEXT_MAIN)
        c.fill = fill(BG_HEADER); c.border = borda_fina()
    ws2.row_dimensions[ultima].height = 22

    # ─── ABA 3: Ranking de Lojas ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Ranking de Lojas")
    ws3.sheet_view.showGridLines = False
    ws3.sheet_properties.tabColor = GREEN

    ws3.merge_cells("A1:G1")
    h3 = ws3["A1"]
    h3.value = f"RANKING DE LOJAS — {mes_fmt}"
    h3.font = fnt(bold=True, size=12, color=TEXT_MAIN); h3.fill = fill(BG_HEADER)
    h3.alignment = aln(); ws3.row_dimensions[1].height = 26

    cols_rank = [
        ("Ranking", 10), ("Loja", 18), ("Receita Bruta", 16),
        ("EBITDA", 14), ("Mg. EBITDA", 12), ("Mg. Bruta", 12), ("Status", 12),
    ]
    aplicar_header(ws3, 2, cols_rank)

    sub2 = sub.copy()
    sub2["mg_ebitda"] = sub2["ebitda"] / sub2["receita_liquida"]
    sub2 = sub2.sort_values("mg_ebitda", ascending=False).reset_index(drop=True)

    for i, row in sub2.iterrows():
        linha = i + 3
        bg = BG_ROW1 if i % 2 == 0 else BG_ROW2
        rank = i + 1
        status = semaforo_ebitda(row["mg_ebitda"])
        status_cor = GREEN if status == "VERDE" else (AMBER if status == "AMARELO" else CORAL)
        vals_r = [rank, row["loja"], row["receita_bruta"], row["ebitda"], row["mg_ebitda"],
                  row["mg_bruta"], status]
        fmts_r = ["num","txt","moeda","moeda","pct","pct","txt"]
        for col, (val, fmt) in enumerate(zip(vals_r, fmts_r), 1):
            c = ws3.cell(linha, col, val)
            c.fill = fill(bg); c.border = borda_fina()
            if fmt == "moeda":
                formatar_valor(c, "moeda")
                c.font = fnt(size=9, color=TEXT_MAIN)
            elif fmt == "pct":
                c.number_format = "0.0%"; c.alignment = aln("right")
                c.font = fnt(size=9, color=TEXT_MAIN)
            elif col == 7:
                c.font = fnt(bold=True, size=9, color=status_cor)
                c.alignment = aln()
            elif col == 1:
                c.font = fnt(bold=True, size=10, color=BLUE); c.alignment = aln()
            else:
                c.font = fnt(size=9, color=TEXT_MAIN); c.alignment = aln("left")
        ws3.row_dimensions[linha].height = 18

    # ─── ABA 4: Inadimplência / Aging ────────────────────────────────────────
    ws4 = wb.create_sheet("Inadimplência Aging")
    ws4.sheet_view.showGridLines = False
    ws4.sheet_properties.tabColor = AMBER

    ws4.merge_cells("A1:J1")
    h4 = ws4["A1"]
    h4.value = f"CARTEIRA DE RECEBÍVEIS — AGING — {mes_fmt}"
    h4.font = fnt(bold=True, size=12, color=TEXT_MAIN); h4.fill = fill(BG_HEADER)
    h4.alignment = aln(); ws4.row_dimensions[1].height = 26

    cols_aging = [
        ("Cliente", 20), ("A Vencer", 14), ("1-30d", 13), ("31-60d", 13),
        ("61-90d", 13), ("91-120d", 13), (">120d", 13),
        ("Total Carteira", 16), ("Total Vencido", 16), ("Taxa Inadim.", 13),
    ]
    aplicar_header(ws4, 2, cols_aging)

    cr_mes = dfs["cr"][dfs["cr"]["mes"] == mes].sort_values("taxa_inadimplencia", ascending=False)
    for i, (_, row) in enumerate(cr_mes.iterrows(), 3):
        bg = BG_ROW1 if i % 2 == 0 else BG_ROW2
        inadim_val = row["taxa_inadimplencia"]
        inadim_cor = CORAL if inadim_val > 0.18 else (AMBER if inadim_val > 0.10 else GREEN)
        vals_cr = [
            row["cliente"], row["a_vencer"], row["venc_1_30"], row["venc_31_60"],
            row["venc_61_90"], row["venc_91_120"], row["venc_acima_120"],
            row["total_carteira"], row["total_vencido"], row["taxa_inadimplencia"],
        ]
        for col, val in enumerate(vals_cr, 1):
            c = ws4.cell(i, col, val)
            c.fill = fill(bg); c.border = borda_fina()
            if col == 1:
                c.font = fnt(size=9, color=TEXT_MAIN); c.alignment = aln("left")
            elif col == 10:
                c.number_format = "0.0%"; c.alignment = aln("right")
                c.font = fnt(bold=True, size=9, color=inadim_cor)
            else:
                formatar_valor(c, "moeda")
                c.font = fnt(size=9, color=TEXT_MAIN)
        ws4.row_dimensions[i].height = 16

    # Totais
    ultima_cr = 3 + len(cr_mes)
    ws4.cell(ultima_cr, 1, "TOTAL").font = fnt(bold=True, size=9, color=TEXT_MAIN)
    ws4.cell(ultima_cr, 1).fill = fill(BG_HEADER); ws4.cell(ultima_cr, 1).border = borda_fina()
    for col_idx, col_let in zip(range(2, 10), ["B","C","D","E","F","G","H","I"]):
        c = ws4.cell(ultima_cr, col_idx, f"=SUM({col_let}3:{col_let}{ultima_cr-1})")
        formatar_valor(c, "moeda")
        c.font = fnt(bold=True, size=9, color=TEXT_MAIN)
        c.fill = fill(BG_HEADER); c.border = borda_fina()
    c_tx = ws4.cell(ultima_cr, 10, f"=I{ultima_cr}/H{ultima_cr}")
    c_tx.number_format = "0.0%"; c_tx.alignment = aln("right")
    c_tx.font = fnt(bold=True, size=9, color=AMBER)
    c_tx.fill = fill(BG_HEADER); c_tx.border = borda_fina()
    ws4.row_dimensions[ultima_cr].height = 22

    # ─── ABA 5: Gráficos ─────────────────────────────────────────────────────
    ws5 = wb.create_sheet("Gráficos")
    ws5.sheet_view.showGridLines = False
    ws5.sheet_properties.tabColor = PURPLE
    ws5.sheet_view.zoomScale = 85

    ws5.merge_cells("A1:P1")
    hg = ws5["A1"]
    hg.value = "DASHBOARD — ANÁLISE VISUAL"
    hg.font = fnt(bold=True, size=12, color=TEXT_MAIN); hg.fill = fill(BG_HEADER)
    hg.alignment = aln(); ws5.row_dimensions[1].height = 26

    graficos = [
        (grafico_receita_ebitda(dfs, mes), 3, 1),
        (grafico_ranking_lojas(dfs, mes), 3, 10),
        (grafico_real_vs_orcado(dfs, mes), 22, 1),
        (grafico_inadimplencia(dfs, mes), 22, 10),
    ]
    for img_bytes, row_start, col_start in graficos:
        img_io = io.BytesIO(img_bytes)
        img = XLImage(img_io)
        img.width  = 490
        img.height = 195
        cell_ref = f"{get_column_letter(col_start)}{row_start}"
        ws5.add_image(img, cell_ref)

    # Ajuste altura das linhas de gráfico
    for r in range(3, 40):
        ws5.row_dimensions[r].height = 13
    for c in range(1, 20):
        ws5.column_dimensions[get_column_letter(c)].width = 7.5

    # ── Salvar ────────────────────────────────────────────────────────────────
    loja_slug = loja.replace(" ", "_").replace("/", "-") if loja != "all" else "consolidado"
    nome = f"relatorio_{mes}_{loja_slug}.xlsx"
    destino = path_out / nome
    wb.save(destino)
    return destino


# ═══════════════════════════════════════════════════════════════════════════════
# ETAPA 4 — Gerador PDF
# ═══════════════════════════════════════════════════════════════════════════════

def gerar_pdf(dfs: dict, kpis: dict, mes: str, loja: str, path_out: Path) -> Path:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (Image as RLImage, PageBreak, Paragraph,
                                    SimpleDocTemplate, Spacer, Table,
                                    TableStyle, HRFlowable)
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

    W, H = A4
    mes_fmt = datetime.strptime(mes, "%Y-%m").strftime("%B de %Y").capitalize()
    loja_slug = loja.replace(" ", "_").replace("/", "-") if loja != "all" else "consolidado"
    nome = f"relatorio_{mes}_{loja_slug}.pdf"
    destino = path_out / nome

    # Cores ReportLab
    def rl_hex(h):
        h = h.lstrip("#")
        return colors.HexColor(f"#{h}")

    BG    = rl_hex("FFFFFF")
    SURF  = rl_hex("F4F6FA")
    BORDR = rl_hex("CBD5E1")
    TXT   = rl_hex("1E293B")
    MUTD  = rl_hex("475569")
    BLUER = rl_hex("1D4ED8")
    TEALR = rl_hex("0D9488")
    GREER = rl_hex("15803D")
    AMBR  = rl_hex("B45309")
    CORLR = rl_hex("B91C1C")
    PURPR = rl_hex("6D28D9")

    doc = SimpleDocTemplate(
        str(destino), pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm,
        topMargin=15*mm, bottomMargin=15*mm,
        title=f"Relatório Financeiro {mes_fmt}",
        author="Rede Máxima Supermercados Ltda.",
    )

    styles = getSampleStyleSheet()

    def estilo(name, parent="Normal", **kw):
        return ParagraphStyle(name, parent=styles[parent], **kw)

    s_titulo   = estilo("Titulo",   fontSize=16, textColor=TXT, alignment=TA_CENTER,
                         spaceAfter=2, fontName="Helvetica-Bold")
    s_subtit   = estilo("SubTit",   fontSize=10, textColor=MUTD, alignment=TA_CENTER,
                         spaceAfter=8, fontName="Helvetica")
    s_secao    = estilo("Secao",    fontSize=11, textColor=BLUER, alignment=TA_LEFT,
                         spaceAfter=4, spaceBefore=10, fontName="Helvetica-Bold")
    s_body     = estilo("Body",     fontSize=8.5, textColor=TXT, alignment=TA_LEFT,
                         spaceAfter=3, leading=13, fontName="Helvetica")
    s_body_r   = estilo("BodyR",    fontSize=8.5, textColor=TXT, alignment=TA_RIGHT,
                         fontName="Helvetica")
    s_kpi_lab  = estilo("KpiLab",   fontSize=7.5, textColor=MUTD, alignment=TA_CENTER,
                         fontName="Helvetica")
    s_kpi_val  = estilo("KpiVal",   fontSize=13, textColor=TXT, alignment=TA_CENTER,
                         fontName="Helvetica-Bold")
    s_nota     = estilo("Nota",     fontSize=7, textColor=MUTD, alignment=TA_LEFT,
                         fontName="Helvetica")

    def fmt_r(v):  return f"R$ {v:,.0f}"
    def fmt_p(v):  return f"{v:.1%}"
    def fmt_n(v):  return f"{v:.0f}"

    def sem_color(tipo, v):
        if tipo == "ebitda": return GREER if v >= 0.20 else (AMBR if v >= 0.15 else CORLR)
        if tipo == "inadim": return GREER if v <= 0.10 else (AMBR if v <= 0.18 else CORLR)
        if tipo == "score":  return GREER if v >= 70 else (AMBR if v >= 50 else CORLR)
        return TXT

    def hr():
        return HRFlowable(width="100%", thickness=0.5, color=BORDR, spaceAfter=4, spaceBefore=4)

    story = []

    # ── Página 1: Capa ───────────────────────────────────────────────────────
    story.append(Spacer(1, 20*mm))
    story.append(Paragraph("REDE MÁXIMA SUPERMERCADOS LTDA.", s_titulo))
    story.append(Paragraph("Pacote de Fechamento Mensal", s_subtit))
    story.append(Spacer(1, 6*mm))

    capa_data = [
        ["Competência:", mes_fmt],
        ["Emissão:", datetime.now().strftime("%d/%m/%Y %H:%M")],
        ["Escopo:", loja if loja != "all" else "Consolidado — todas as lojas"],
        ["Gerado por:", "Sistema de Automação — scripts/gerar_relatorio.py"],
    ]
    t_capa = Table(capa_data, colWidths=[45*mm, 120*mm])
    t_capa.setStyle(TableStyle([
        ("TEXTCOLOR", (0,0), (-1,-1), TXT),
        ("FONTNAME",  (0,0), (0,-1), "Helvetica-Bold"),
        ("FONTNAME",  (1,0), (1,-1), "Helvetica"),
        ("FONTSIZE",  (0,0), (-1,-1), 9),
        ("ROWBACKGROUNDS", (0,0), (-1,-1), [rl_hex("FFFFFF"), rl_hex("F4F6FA")]),
        ("GRID",      (0,0), (-1,-1), 0.4, BORDR),
        ("TOPPADDING",(0,0),(-1,-1), 5),
        ("BOTTOMPADDING",(0,0),(-1,-1), 5),
    ]))
    story.append(t_capa)
    story.append(Spacer(1, 8*mm))
    story.append(hr())

    # ── Página 2: KPIs + Comentário ──────────────────────────────────────────
    story.append(Paragraph("KPIs Executivos", s_secao))

    kpi_tab_data = [
        ["KPI", "Valor", "Meta", "Status"],
        ["Receita Bruta",    fmt_r(kpis["receita_bruta"]),    "—",     "—"],
        ["Receita Líquida",  fmt_r(kpis["receita_liquida"]), "—",     "—"],
        ["EBITDA",           fmt_r(kpis["ebitda"]),          "—",     "—"],
        ["Margem EBITDA",    fmt_p(kpis["margem_ebitda"]),   "≥ 20%", semaforo_ebitda(kpis["margem_ebitda"])],
        ["Margem Bruta",     fmt_p(kpis["margem_bruta"]),    "—",     "—"],
        ["Margem Líquida",   fmt_p(kpis["margem_liquida"]),  "—",     "—"],
        ["Inadimplência",    fmt_p(kpis["inadimplencia"]),   "≤ 10%", semaforo_inadim(kpis["inadimplencia"])],
        ["FCO",              fmt_r(kpis["fco"]),             "—",     "—"],
        ["Saldo Final Caixa",fmt_r(kpis["saldo_final"]),     "—",     "—"],
        ["Score Saúde",      fmt_n(kpis["score_saude"]),     "≥ 70",  semaforo_score(kpis["score_saude"])],
        ["Var. Receita MoM", fmt_p(kpis["var_receita_mom"]), "—",     "↑" if kpis["var_receita_mom"] >= 0 else "↓"],
        ["Var. EBITDA vs Orç.", fmt_p(kpis["var_ebitda_orc"]),"—",    "↑" if kpis["var_ebitda_orc"] >= 0 else "↓"],
    ]

    def _cor_status(txt):
        if txt == "VERDE": return GREER
        if txt == "AMARELO": return AMBR
        if txt == "VERMELHO": return CORLR
        if txt == "↑": return GREER
        if txt == "↓": return CORLR
        return MUTD

    ts_kpi = [
        ("FONTNAME",  (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",  (0,0), (-1,-1), 8.5),
        ("TEXTCOLOR", (0,0), (-1,0),  TXT),
        ("BACKGROUND",(0,0), (-1,0),  rl_hex("F4F6FA")),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [rl_hex("FFFFFF"), rl_hex("F4F6FA")]),
        ("GRID", (0,0), (-1,-1), 0.4, BORDR),
        ("TEXTCOLOR", (0,1), (0,-1), MUTD),
        ("TEXTCOLOR", (1,1), (1,-1), TXT),
        ("TEXTCOLOR", (2,1), (2,-1), MUTD),
        ("ALIGN", (1,0), (-1,-1), "RIGHT"),
        ("ALIGN", (0,0), (0,-1), "LEFT"),
        ("TOPPADDING",    (0,0),(-1,-1), 4),
        ("BOTTOMPADDING", (0,0),(-1,-1), 4),
    ]
    for row_i, row_d in enumerate(kpi_tab_data[1:], 1):
        txt_status = row_d[3]
        ts_kpi.append(("TEXTCOLOR", (3, row_i), (3, row_i), _cor_status(txt_status)))
        ts_kpi.append(("FONTNAME",  (3, row_i), (3, row_i), "Helvetica-Bold"))

    t_kpis = Table(kpi_tab_data, colWidths=[55*mm, 42*mm, 30*mm, 38*mm])
    t_kpis.setStyle(TableStyle(ts_kpi))
    story.append(t_kpis)
    story.append(Spacer(1, 5*mm))
    story.append(hr())

    story.append(Paragraph("Comentário Executivo", s_secao))
    for linha_txt in gerar_comentario(kpis):
        story.append(Paragraph(linha_txt, s_body))
    story.append(Spacer(1, 5*mm))
    story.append(hr())

    # ── Página 3: Gráficos ───────────────────────────────────────────────────
    story.append(PageBreak())
    story.append(Paragraph("Análise Visual", s_secao))

    for img_bytes, titulo in [
        (grafico_receita_ebitda(dfs, mes), "Receita Líquida vs EBITDA — últimos 12 meses"),
        (grafico_ranking_lojas(dfs, mes),  "Ranking de Lojas — Margem EBITDA"),
        (grafico_real_vs_orcado(dfs, mes), "Real vs Orçado por Loja — Receita Bruta"),
        (grafico_inadimplencia(dfs, mes),  "Taxa de Inadimplência — últimos 12 meses"),
    ]:
        story.append(Paragraph(titulo, s_nota))
        img_io = io.BytesIO(img_bytes)
        img_rl = RLImage(img_io, width=165*mm, height=55*mm)
        story.append(img_rl)
        story.append(Spacer(1, 3*mm))

    story.append(hr())

    # ── Página 4: Tabela DRE consolidada ────────────────────────────────────
    story.append(PageBreak())
    story.append(Paragraph("DRE Consolidada por Loja", s_secao))

    sub = dfs["dre"][dfs["dre"]["mes"] == mes].copy()
    if loja != "all": sub = sub[sub["loja"] == loja]
    sub["mg_bruta"]  = sub["lucro_bruto"] / sub["receita_liquida"]
    sub["mg_ebitda"] = sub["ebitda"] / sub["receita_liquida"]

    dre_tab = [["Loja", "Rec. Líquida", "CMV", "Lucro Bruto", "EBITDA", "Mg. EBITDA"]]
    for _, row in sub.iterrows():
        dre_tab.append([
            row["loja"].replace("L0","L"),
            fmt_r(row["receita_liquida"]),
            fmt_r(row["cmv"]),
            fmt_r(row["lucro_bruto"]),
            fmt_r(row["ebitda"]),
            fmt_p(row["mg_ebitda"]),
        ])
    # Totais
    dre_tab.append([
        "TOTAL",
        fmt_r(sub["receita_liquida"].sum()),
        fmt_r(sub["cmv"].sum()),
        fmt_r(sub["lucro_bruto"].sum()),
        fmt_r(sub["ebitda"].sum()),
        fmt_p(sub["ebitda"].sum() / sub["receita_liquida"].sum()),
    ])

    ts_dre = [
        ("FONTNAME",  (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",  (0,0), (-1,-1), 8),
        ("TEXTCOLOR", (0,0), (-1,0),  TXT),
        ("BACKGROUND",(0,0), (-1,0),  SURF),
        ("ROWBACKGROUNDS", (0,1), (-1,-2), [rl_hex("FFFFFF"), rl_hex("F4F6FA")]),
        ("BACKGROUND",(0,-1),(-1,-1), SURF),
        ("FONTNAME",  (0,-1),(-1,-1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0,1), (-1,-1), TXT),
        ("GRID", (0,0), (-1,-1), 0.4, BORDR),
        ("ALIGN", (1,0), (-1,-1), "RIGHT"),
        ("ALIGN", (0,0), (0,-1), "LEFT"),
        ("TOPPADDING",    (0,0),(-1,-1), 4),
        ("BOTTOMPADDING", (0,0),(-1,-1), 4),
    ]
    t_dre = Table(dre_tab, colWidths=[42*mm, 33*mm, 28*mm, 30*mm, 28*mm, 24*mm])
    t_dre.setStyle(TableStyle(ts_dre))
    story.append(t_dre)

    story.append(Spacer(1, 5*mm))
    story.append(hr())
    story.append(Paragraph(
        f"Documento gerado automaticamente em {datetime.now():%d/%m/%Y %H:%M} | "
        "Rede Máxima Supermercados Ltda. | Uso interno",
        s_nota
    ))

    doc.build(story)
    return destino


# ═══════════════════════════════════════════════════════════════════════════════
# ETAPA 5 — CLI + orquestrador
# ═══════════════════════════════════════════════════════════════════════════════

def _validar_mes(valor: str) -> str:
    try:
        datetime.strptime(valor, "%Y-%m")
    except ValueError:
        raise argparse.ArgumentTypeError(f"Formato de mês inválido: '{valor}'. Use YYYY-MM.")
    return valor


def main():
    parser = argparse.ArgumentParser(
        description="Automação de Relatório Financeiro — Rede Máxima Supermercados",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemplos:
  python scripts/gerar_relatorio.py --mes 2024-06
  python scripts/gerar_relatorio.py --mes 2024-06 --loja L01-Centro
  python scripts/gerar_relatorio.py --mes 2024-06 --formato excel
  python scripts/gerar_relatorio.py --mes 2024-06 --formato pdf
  python scripts/gerar_relatorio.py --mes 2024-06 --formato excel,pdf --output /tmp
        """,
    )
    parser.add_argument("--mes",     type=_validar_mes, required=True,
                        help="Mês de competência no formato YYYY-MM")
    parser.add_argument("--loja",    default="all",
                        help="Código da loja (ex: L01-Centro) ou 'all' para consolidado")
    parser.add_argument("--formato", default="excel,pdf",
                        help="Formatos de saída: excel, pdf ou excel,pdf (padrão)")
    parser.add_argument("--output",  default=str(OUT),
                        help=f"Diretório de saída (padrão: {OUT})")
    args = parser.parse_args()

    formatos = [f.strip().lower() for f in args.formato.split(",")]
    validos  = {"excel", "pdf"}
    invalidos = set(formatos) - validos
    if invalidos:
        print(f"Erro: formato(s) inválido(s): {invalidos}. Use: excel, pdf ou excel,pdf")
        sys.exit(1)

    path_out = Path(args.output)
    path_out.mkdir(parents=True, exist_ok=True)

    print(f"\n{'='*60}")
    print(f"  REDE MÁXIMA SUPERMERCADOS — FECHAMENTO MENSAL")
    print(f"  Mês: {args.mes}  |  Loja: {args.loja}  |  Formato: {args.formato}")
    print(f"{'='*60}")

    print("  Carregando dados...", end=" ", flush=True)
    dfs = carregar_dados()
    meses_disp = sorted(dfs["dre"]["mes"].unique())
    if args.mes not in meses_disp:
        print(f"\nErro: mês '{args.mes}' não encontrado nos dados.")
        print(f"Meses disponíveis: {meses_disp[0]} a {meses_disp[-1]}")
        sys.exit(1)
    print("OK")

    print("  Calculando KPIs...", end=" ", flush=True)
    kpis = calcular_kpis(dfs, args.mes, args.loja)
    print("OK")

    arquivos = []

    if "excel" in formatos:
        print("  Gerando Excel...", end=" ", flush=True)
        p = gerar_excel(dfs, kpis, args.mes, args.loja, path_out)
        arquivos.append(p)
        print(f"OK → {p.name}")

    if "pdf" in formatos:
        print("  Gerando PDF...", end=" ", flush=True)
        p = gerar_pdf(dfs, kpis, args.mes, args.loja, path_out)
        arquivos.append(p)
        print(f"OK → {p.name}")

    print(f"\n  Score de Saúde Financeira: {kpis['score_saude']:.0f}/100 "
          f"({semaforo_score(kpis['score_saude'])})")
    print(f"  Margem EBITDA:            {kpis['margem_ebitda']:.1%} "
          f"({semaforo_ebitda(kpis['margem_ebitda'])})")
    print(f"  Inadimplência:            {kpis['inadimplencia']:.1%} "
          f"({semaforo_inadim(kpis['inadimplencia'])})")
    print(f"\n  Arquivos gerados em: {path_out}")
    for a in arquivos:
        print(f"    • {a.name}  ({a.stat().st_size/1024:.0f} KB)")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
